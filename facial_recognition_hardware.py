import face_recognition
import cv2
import numpy as np
from picamera2 import Picamera2
import time
import pickle
from gpiozero import LED
import json
from datetime import datetime
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# 1. Google Sheets connection
# ─────────────────────────────────────────────
scope = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
creds = ServiceAccountCredentials.from_json_keyfile_name("service_account.json", scope)
client = gspread.authorize(creds)
sheet = client.open("Attendance Log").sheet1

# Load roster from Google Sheets
def load_roster():
    roster = {}
    records = sheet.get_all_values()
    for i, row in enumerate(records[1:], start=2):
        if len(row) >= 2:
            name = row[0].strip()
            student_id = row[1].strip()
            if name:
                roster[name] = {'row': i, 'id': student_id, 'name': name}
    return roster

roster = load_roster()
print(f"[INFO] Loaded {len(roster)} students: {[v['name'] for v in roster.values()]}")

# Reset all students to Unknown at start of session
def initialise_sheet_for_today():
    print("[INFO] Resetting sheet: marking all students as Unknown...")
    for student in roster.values():
        row = student['row']
        try:
            sheet.update(f'C{row}:D{row}', [["Unknown", ""]])
        except Exception as e:
            print(f"[ERROR] Could not reset row {row} for {student['name']}: {e}")
    print("[INFO] Sheet reset complete.")

initialise_sheet_for_today()

today_str = datetime.now().strftime("%Y-%m-%d")
local_log_file = f"attendance_{today_str}.json"
logged_times = {}

# ─────────────────────────────────────────────
# 2. Load face encodings
# ─────────────────────────────────────────────
print("[INFO] loading encodings...")
with open("encodings.pickle", "rb") as f:
    data = pickle.loads(f.read())
known_face_encodings = data["encodings"]
known_face_names = data["names"]

# ─────────────────────────────────────────────
# 3. Camera and GPIO
# ─────────────────────────────────────────────
# Initialize the camera
picam2 = Picamera2()
picam2.configure(picam2.create_preview_configuration(main={"format": 'XRGB8888', "size": (640, 480)})) # lower resolution from 1920x1080 to 640x480
picam2.start()

# Initialize GPIO
output = LED(14)

# ─────────────────────────────────────────────
# 4. State variables
# ─────────────────────────────────────────────
cv_scaler = 8 # changed from 4 to 8 to reduce mem strain and boost FPS

face_locations = []
face_encodings = []
face_names = []
frame_count = 0
start_time = time.time()
fps = 0

# Keeps track of who was logged this session (no duplicate entries)
already_logged = []

# Tracks names detected in the CURRENT frame (for live sidebar display)
current_frame_names = []

# Running counters for the teacher window
recognized_count = 0
unrecognized_count = 0
# Full session log: list of dicts {name, status, time}
session_log = []

# List of names that will trigger the GPIO pin
authorized_names = ["Marc", "John", "Soren", "Ryan", "Paul", "Nick"]

# ─────────────────────────────────────────────
# 5. Window names
# ─────────────────────────────────────────────
WINDOW_STUDENT = 'Student View'
WINDOW_TEACHER = 'Teacher View (Admin)'

# ─────────────────────────────────────────────
# 6. Core processing
# ─────────────────────────────────────────────
def process_frame(frame):
    global face_locations, face_encodings, face_names
    global already_logged, logged_times
    global recognized_count, unrecognized_count, session_log
    global current_frame_names
    
    # Resize the frame using cv_scaler to increase performance (less pixels processed, less time spent)
    resized_frame = cv2.resize(frame, (0, 0), fx=(1/cv_scaler), fy=(1/cv_scaler))
    
    # Convert the image from BGR to RGB colour space, the facial recognition library uses RGB, OpenCV uses BGR
    rgb_resized_frame = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
    
    # Find all the faces and face encodings in the current frame of video
    face_locations = face_recognition.face_locations(rgb_resized_frame)
    face_encodings = face_recognition.face_encodings(rgb_resized_frame, face_locations, model='large')
    
    face_names = []
    current_frame_names = []
    authorized_face_detected = False
    
    for face_encoding in face_encodings:
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
        
        # Use the known face with the smallest distance to the new face
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]
            
            #JSON logic for creating JSON log for face that are recognized
            
            #CHECK 1: Is this a known student?
            #Check 2: have we already logged them in this session?
            if name != "Unknown" and name not in already_logged:
                
                # create timestamp variable (to match time between local and cloud data)
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                time_only = datetime.now().strftime("%H:%M:%S")  # time-only string for the sheet's Time column
                
                # check the student exists in the roster we loaded from Google Sheets at startup
                if name in roster:  # removed name_lower, now compares directly against roster
                    row = roster[name]['row']  # get the exact sheet row number for this student
                    try:
                        # updates the student's existing row instead of appending a new row at the bottom
                        sheet.update(f'C{row}:D{row}', [["Present", time_only]])
                        print(f"[CLOUD] Updated row {row} → {name}: Present at {time_only}")
                    except Exception as e:
                        print(f"[ERROR] Could not update Google Sheet for {name}: {e}")
                else:
                    print(f"[WARNING] {name} recognised but not found in roster.")  # NEW: warns if face is known but missing from the sheet

                # add the name to our 'memory' list so they are not logged again
                already_logged.append(name)  # Prevents duplicates
                logged_times[name] = timestamp  # NEW: store timestamp so export_to_xlsx() can reference it when R is pressed

                # replaces the old append logic — instead of adding one line per detection,
                # we rewrite the entire file each time as a clean snapshot (one entry per student)
                snapshot = []
                for student in roster.values():
                    n = student['name']
                    snapshot.append({
                        "student_name": n,
                        "student_id": student['id'],  # includes student ID
                        "status": "Present" if n in already_logged else "Unknown",  # default is "Unknown" instead of "Absent"
                        "time_logged": logged_times.get(n, ""),
                        "date": today_str
                    })
                # "w" overwrites the file each time instead of "a" which appended
                with open(local_log_file, "w") as f:
                    json.dump(snapshot, f, indent=2)
                print(f"[LOCAL] Snapshot updated: {name} marked Present at {timestamp}")

                recognized_count += 1
                session_log.append({
                    "name": name,
                    "id": roster[name]['id'] if name in roster else "N/A",
                    "status": "Recognized",
                    "time": time_only
                })
                
            # --- END OF JSON logic block ---

        # ── Track unknown faces that are genuinely new ──
        # (only count/log them once per session via a separate set)
        if name == "Unknown":
            # We don't add to already_logged for unknowns, but we track
            # them in the session log once per detection run to avoid spam.
            # Use a simple cooldown: only log if not in session_log recently.
            recent_unknowns = [e for e in session_log if e["status"] == "Unrecognized"]
            # Simple dedup: add at most one "Unknown" entry per 10 seconds
            if not recent_unknowns or (
                datetime.now() - datetime.strptime(
                    recent_unknowns[-1]["time"], "%H:%M:%S"
                )
            ).seconds > 10:
                time_only = datetime.now().strftime("%H:%M:%S")
                unrecognized_count += 1
                session_log.append({
                    "name": "Unknown",
                    "id": "—",
                    "status": "Unrecognized",
                    "time": time_only
                })
            
        # Check if the detected face is in our authorized list
        if name in authorized_names:
            authorized_face_detected = True

        face_names.append(name)
        current_frame_names.append(name)
    
    # ============================================
    # Control the GPIO pin based on face detection
    # ============================================
    if authorized_face_detected:
        output.on()  # Turn on Pin
    else:
        output.off()  # Turn off Pin
    
    return frame

# ─────────────────────────────────────────────
# 7. Draw both views
# ─────────────────────────────────────────────
def draw_results(frame):
    """
    Returns two frames:
      student_view  – camera feed (left) + info panel (right)
      teacher_view  – black background with session log and counters
    """
    h, w = frame.shape[:2]

    # ── Colours ──
    GREEN       = (0, 220, 80)
    RED         = (50, 50, 220)
    WHITE       = (255, 255, 255)
    DARK_GRAY   = (35, 35, 35)
    BLACK       = (0, 0, 0)
    PANEL_BG    = (20, 20, 20)
    GREEN_BG    = (30, 110, 50)
    RED_BG      = (50, 30, 160)
    GREEN_TEXT  = (100, 255, 140)
    RED_TEXT    = (130, 130, 255)
    MUTED       = (160, 160, 160)
    FONT        = cv2.FONT_HERSHEY_DUPLEX
    FONT_SIMPLE = cv2.FONT_HERSHEY_SIMPLEX

    # ════════════════════════════════
    # STUDENT VIEW
    # camera feed (left) + side panel (right)
    # ════════════════════════════════
    PANEL_W = 280   # width of the right info panel
    student_view = np.zeros((h, w + PANEL_W, 3), dtype=np.uint8)

    # Paste the camera frame on the left
    student_view[:, :w] = frame

    # Draw face bounding boxes on the camera side
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top    *= cv_scaler
        right  *= cv_scaler
        bottom *= cv_scaler
        left   *= cv_scaler

        colour = GREEN if name != "Unknown" else RED

        cv2.rectangle(student_view, (left, top), (right, bottom), colour, 2)
        cv2.rectangle(student_view, (left, bottom - 30), (right, bottom), colour, cv2.FILLED)
        cv2.putText(student_view, name, (left + 6, bottom - 8),
                    FONT, 0.75, WHITE, 1)

    # ── Right info panel background ──
    student_view[:, w:] = PANEL_BG

    # Pick the "most recent" recognised face to feature in the panel.
    # If multiple faces are in frame, prefer the first recognised one.
    featured_name = None
    featured_id   = "—"
    is_recognised = False

    for name in current_frame_names:
        if name != "Unknown":
            featured_name  = name
            featured_id    = roster[name]['id'] if name in roster else "N/A"
            is_recognised  = True
            break
    if featured_name is None and current_frame_names:
        featured_name = "Unknown"

    # ── Student info block (top of panel) ──
    info_x = w + 18
    cv2.putText(student_view, "Student Info", (info_x, 36),
                FONT_SIMPLE, 0.65, MUTED, 1)
    cv2.line(student_view, (w + 10, 46), (w + PANEL_W - 10, 46), (60, 60, 60), 1)

    if featured_name:
        cv2.putText(student_view, featured_name, (info_x, 82),
                    FONT, 0.9, WHITE, 1)
        cv2.putText(student_view, f"ID: {featured_id}", (info_x, 114),
                    FONT_SIMPLE, 0.6, MUTED, 1)
    else:
        cv2.putText(student_view, "No face detected", (info_x, 82),
                    FONT_SIMPLE, 0.6, MUTED, 1)

    # ── Status banner (bottom 30% of panel) ──
    banner_top = int(h * 0.70)
    cv2.line(student_view, (w + 10, banner_top - 1),
             (w + PANEL_W - 10, banner_top - 1), (60, 60, 60), 1)

    if current_frame_names:
        if is_recognised:
            cv2.rectangle(student_view,
                          (w, banner_top), (w + PANEL_W, h),
                          GREEN_BG, cv2.FILLED)
            cv2.putText(student_view, "Welcome!", (info_x, banner_top + 50),
                        FONT, 0.95, GREEN_TEXT, 2)
            cv2.putText(student_view, "Attendance logged", (info_x, banner_top + 82),
                        FONT_SIMPLE, 0.55, GREEN_TEXT, 1)
        else:
            cv2.rectangle(student_view,
                          (w, banner_top), (w + PANEL_W, h),
                          RED_BG, cv2.FILLED)
            cv2.putText(student_view, "Unrecognized", (info_x, banner_top + 44),
                        FONT, 0.85, RED_TEXT, 2)
            cv2.putText(student_view, "Please see a TA", (info_x, banner_top + 74),
                        FONT_SIMPLE, 0.55, RED_TEXT, 1)
            cv2.putText(student_view, "or Professor", (info_x, banner_top + 96),
                        FONT_SIMPLE, 0.55, RED_TEXT, 1)
    else:
        # No one in frame – neutral grey panel
        cv2.rectangle(student_view,
                      (w, banner_top), (w + PANEL_W, h),
                      (40, 40, 40), cv2.FILLED)
        cv2.putText(student_view, "Awaiting student...", (info_x, banner_top + 60),
                    FONT_SIMPLE, 0.55, MUTED, 1)

    # ════════════════════════════════
    # TEACHER / ADMIN VIEW
    # Black background, no camera feed
    # ════════════════════════════════
    TV_W = 520
    TV_H = h
    teacher_view = np.zeros((TV_H, TV_W, 3), dtype=np.uint8)

    # Title
    cv2.putText(teacher_view, "Attendance Session", (20, 42),
                FONT, 0.9, WHITE, 1)
    cv2.line(teacher_view, (10, 56), (TV_W - 10, 56), (70, 70, 70), 1)

    # ── Counter cards ──
    card_y  = 72
    card_h  = 60
    card_w  = (TV_W - 50) // 2   # two cards side by side with a gap

    # Recognised card
    cv2.rectangle(teacher_view, (15, card_y),
                  (15 + card_w, card_y + card_h), (25, 65, 35), cv2.FILLED)
    cv2.rectangle(teacher_view, (15, card_y),
                  (15 + card_w, card_y + card_h), (50, 140, 70), 1)
    cv2.putText(teacher_view, "Recognized", (25, card_y + 22),
                FONT_SIMPLE, 0.52, GREEN_TEXT, 1)
    cv2.putText(teacher_view, str(recognized_count), (25, card_y + 50),
                FONT, 0.95, GREEN_TEXT, 2)

    # Unrecognised card
    cx2 = 20 + card_w + 15
    cv2.rectangle(teacher_view, (cx2, card_y),
                  (cx2 + card_w, card_y + card_h), (65, 25, 35), cv2.FILLED)
    cv2.rectangle(teacher_view, (cx2, card_y),
                  (cx2 + card_w, card_y + card_h), (140, 50, 70), 1)
    cv2.putText(teacher_view, "Unrecognized", (cx2 + 10, card_y + 22),
                FONT_SIMPLE, 0.52, RED_TEXT, 1)
    cv2.putText(teacher_view, str(unrecognized_count), (cx2 + 10, card_y + 50),
                FONT, 0.95, RED_TEXT, 2)

    # ── Session log ──
    log_start_y = card_y + card_h + 28
    cv2.putText(teacher_view, "Session Log", (15, log_start_y),
                FONT_SIMPLE, 0.55, MUTED, 1)
    cv2.line(teacher_view, (10, log_start_y + 8),
             (TV_W - 10, log_start_y + 8), (55, 55, 55), 1)

    max_visible = (TV_H - log_start_y - 80) // 28  # how many rows fit
    visible_log = session_log[-max_visible:] if len(session_log) > max_visible else session_log

    y_log = log_start_y + 30
    for entry in visible_log:
        dot_colour  = GREEN if entry["status"] == "Recognized" else RED
        name_colour = GREEN_TEXT if entry["status"] == "Recognized" else RED_TEXT
        cv2.circle(teacher_view, (20, y_log - 5), 5, dot_colour, cv2.FILLED)
        label = f"{entry['name']}  {entry['time']}"
        cv2.putText(teacher_view, label, (34, y_log),
                    FONT_SIMPLE, 0.52, name_colour, 1)
        y_log += 28

    # ── Bottom hint bar ──
    hint_y = TV_H - 36
    cv2.rectangle(teacher_view, (0, hint_y - 10),
                  (TV_W, TV_H), (25, 25, 25), cv2.FILLED)
    cv2.line(teacher_view, (10, hint_y - 11),
             (TV_W - 10, hint_y - 11), (60, 60, 60), 1)
    cv2.putText(teacher_view, "R: Export Report    Q: Quit",
                (20, hint_y + 14), FONT_SIMPLE, 0.52, MUTED, 1)

    return student_view, teacher_view

# ─────────────────────────────────────────────
# 8. XLSX export  (unchanged from original)
# ─────────────────────────────────────────────
def export_to_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="2E75B6")
    present_fill = PatternFill("solid", start_color="C6EFCE")
    absent_fill  = PatternFill("solid", start_color="FFCCCC")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers    = ["Name", "ID", "Status", "Time"]
    col_widths = [20, 10, 12, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, student in enumerate(roster.values(), start=2):
        name       = student['name']
        student_id = student['id']
        if name in already_logged:
            status, time_val, row_fill = "Present", logged_times.get(name, ""), present_fill
        else:
            status, time_val, row_fill = "Absent", "", absent_fill

        for col, (val, align) in enumerate(zip([name, student_id, status, time_val],
                                                [left_align, center_align, center_align, center_align]), start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = align
            cell.border = border
            cell.font = Font(name="Arial", size=10)
        ws.row_dimensions[row_idx].height = 18

    filename = f"attendance_report_{today_str}.xlsx"
    wb.save(filename)
    print(f"[EXPORT] Spreadsheet saved as '{filename}'")

# ─────────────────────────────────────────────
# 9. FPS helper
# ─────────────────────────────────────────────
def calculate_fps():
    global frame_count, start_time, fps
    frame_count += 1
    elapsed_time = time.time() - start_time
    if elapsed_time > 1:
        fps = frame_count / elapsed_time
        frame_count = 0
        start_time = time.time()
    return fps

# ─────────────────────────────────────────────
# 10. Window setup
# ─────────────────────────────────────────────
cv2.namedWindow(WINDOW_STUDENT, cv2.WINDOW_NORMAL)
cv2.namedWindow(WINDOW_TEACHER, cv2.WINDOW_NORMAL)

# ─────────────────────────────────────────────
# 11. Main loop
# ─────────────────────────────────────────────
running = True

while True:
    # Capture a frame from camera
    frame = picam2.capture_array()
    frame = cv2.cvtColor(frame, cv2.COLOR_BGRA2BGR)
    
    # Process the frame with the function
    processed_frame = process_frame(frame)
    
    # Get the text and boxes to be drawn based on the processed frame
    display_student, display_teacher = draw_results(processed_frame)
    
    # Calculate and update FPS
    current_fps = calculate_fps()
    
    # Attach FPS counter to the text and boxes
    cv2.putText(display_student, f"FPS: {current_fps:.1f}",
                (10, display_student.shape[0] - 12),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (80, 80, 80), 1)
    
    # Display everything over the video feed.
    cv2.imshow(WINDOW_STUDENT, display_student)
    cv2.imshow(WINDOW_TEACHER, display_teacher)

    # UPDATED key handling
    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        break
    elif key == ord("r"):
        export_to_xlsx()

# ─────────────────────────────────────────────
# 12. Cleanup
# ─────────────────────────────────────────────
# By breaking the loop we run this code here which closes everything
cv2.destroyAllWindows()
picam2.stop()
output.off()  # Make sure to turn off the GPIO pin when exiting
